"""
populate_testcycle.py
─────────────────────
Updates ALL columns in testcycle from my_master.xlsx, including move sequences.

Columns updated (by position in the 49-column sheet):
  [1]  Current Day Cycle
  [2]  Number of Cars
  [3]  Revenue Runs
  [4]  Total Deadhead Runs
  [5]  Total Runs
  [6]  Starting Outlying Point
  [7]  First Move
  [8–43] Move sequences: Train ID / Start / End / Dwell × 9   ← REBUILT from master
  [44] Ending Outlying Point
  [45] Final Move
  [46] Next day cycle
  [47] CMF Train?
  [48] CMF Arrival Time
  [49] CMF Departure Time

Move sequence logic:
  OLP cycles  (layover ≠ CMF, has midday CMF dwell):
      morning trains → last_morningQ (LAUS→CMF) → first_eveningQ (CMF→LAUS) → evening trains
  OLP cycles  (layover ≠ CMF, no CMF):
      morning trains → evening trains
  CMF cycles  (layover = CMF):
      first_morningQ (CMF→LAUS) → morning trains → evening trains → last_eveningQ (LAUS→CMF)

Usage:
  python populate_testcycle.py
      --master   my_master.xlsx
      --template testcycle.xlsx
      --output   testcycle_updated.xlsx
"""

import argparse, re, math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ───────────────────────────────────────────────────────────────────
def na(v):
    if v is None: return ""
    try:
        if math.isnan(float(v)): return ""
    except: pass
    return v

def clean(v):
    s = str(na(v)).strip()
    return "" if s.lower() in ("nan","none","nat") else s

def to_int(v):
    s = clean(v)
    if s.replace('.','').isdigit():
        return int(float(s))
    return ""

def parse_trains(raw):
    """Parse a comma-separated train ID string into a list of ints."""
    s = clean(raw)
    if not s or s.upper() in ('LAYOVER', 'XXX', 'NAN', 'NONE'):
        return []
    return [int(t.strip()) for t in s.split(',') if t.strip().isdigit()]

def calc_time_diff(t1_str, t2_str):
    """
    Return 'X hr Y min' string for (t2 - t1).
    Handles overnight wrap-around.  Returns '' on error.
    """
    t1 = clean(t1_str); t2 = clean(t2_str)
    if not t1 or not t2: return ''
    if t1.startswith('Turn') or t2.startswith('Turn'): return 'Turn'
    if t1 == 'LAYOVER' or t2 == 'LAYOVER': return 'LAYOVER'
    try:
        h1, m1 = map(int, t1.split(':'))
        h2, m2 = map(int, t2.split(':'))
        diff = (h2*60+m2) - (h1*60+m1)
        if diff < 0: diff += 1440          # overnight
        h, m = divmod(diff, 60)
        if h == 0: return f'{m} min'
        if m == 0: return f'{h} hr'
        return f'{h} hr {m} min'
    except:
        return ''

# ── column positions (1-based) ────────────────────────────────────────────────
COL_CYCLE      = 1
COL_CARS       = 2
COL_REV        = 3
COL_DH         = 4
COL_TOT        = 5
COL_START      = 6
COL_FIRST_MOVE = 7
# cols 8-43 = move sequences (4 cols × 9 moves)
COL_END        = 44
COL_FINAL_MOVE = 45
COL_NEXT       = 46
COL_CMF        = 47
COL_CMF_ARR    = 48
COL_CMF_DEP    = 49

MOVE_BASE      = 8          # first move column
MOVE_STRIDE    = 4          # Train ID, Start, End, Dwell
MAX_MOVES      = 9

HEADER_COLS = {COL_CYCLE, COL_CARS, COL_REV, COL_DH, COL_TOT,
               COL_START, COL_FIRST_MOVE, COL_END, COL_FINAL_MOVE,
               COL_NEXT, COL_CMF, COL_CMF_ARR, COL_CMF_DEP}

# ── build lookup tables from master ──────────────────────────────────────────
def build_lookups(master_path):
    df_cw  = pd.read_excel(master_path, sheet_name='cycles_weekday')
    df_cwk = pd.read_excel(master_path, sheet_name='cycles_weekend')
    df_tr  = pd.read_excel(master_path, sheet_name='train_reference')
    df_ld  = pd.read_excel(master_path, sheet_name='laus_dwell')
    df_od  = pd.read_excel(master_path, sheet_name='olp_dwell')

    # ── train routes: train_id(int) → (from_station, to_station) ─────────────
    # Include both Weekday and Weekend/Holiday trains for Sa/Su move sequences
    train_routes = {}
    for _, r in df_tr.iterrows():
        stype = clean(r['service_type'])
        if stype in ('Weekday', 'Weekend/Holiday'):
            try:
                tid = int(float(r['train_id']))
                frm = clean(r['from_station'])
                to  = clean(r['to_station'])
                train_routes[tid] = (frm, to)
            except: pass

    # ── LAUS dwell: (ib_train_str, ob_train_str) → dwell_str ─────────────────
    laus_dwell_lkp = {}
    for _, r in df_ld.iterrows():
        if clean(r.get('day_type','')) == 'Weekday':
            ib    = clean(r['ib_train'])
            ob    = clean(r['ob_train'])
            dwell = clean(r['dwell'])
            if ib and ob and dwell:
                laus_dwell_lkp[(ib, ob)] = dwell

    # ── OLP dwell: (ob_train_str, ib_train_str) → dwell_str ──────────────────
    # ob_train arrives at OLP; ib_train departs from OLP
    olp_dwell_lkp = {}
    for _, r in df_od.iterrows():
        if clean(r.get('day_type','')) == 'Weekday':
            ob    = clean(r['ob_train'])
            ib    = clean(r['ib_train'])
            dwell = clean(r['dwell'])
            if ob and ib and dwell and ob not in ('O/B SYM','') and ib not in ('I/B SYM',''):
                olp_dwell_lkp[(ob, ib)] = dwell

    # ── weekday cycle data ────────────────────────────────────────────────────
    wd = {}
    for _, r in df_cw.iterrows():
        c = na(r['cycle'])
        if c == "": continue
        cnum = int(float(c))

        cars      = na(r['cars'])
        raw_trips = clean(r['weekday_trips'])

        if cars == "":
            continue  # suspended

        rev = 'TBD' if raw_trips.upper() == 'TBD' else to_int(raw_trips)

        cmf_arr = clean(r['cmf_arrive_time'])
        cmf_dep = clean(r['cmf_depart_time'])
        if cmf_arr == "00:01": cmf_arr = "LAYOVER"
        if cmf_dep == "00:01": cmf_dep = "LAYOVER"

        cmf = "Yes" if (cmf_arr or cmf_dep) else "No"
        dh  = 2 if cmf == "Yes" else 0
        tot = (rev if isinstance(rev, int) else 0) + dh

        first   = clean(r['depart_layover']) or clean(r['depart_cmf'])
        final   = clean(r['arrive_layover']) or clean(r['arrive_cmf'])
        layover = clean(r['layover_location'])

        morning_raw = clean(r['morning_trains'])
        evening_raw = clean(r['evening_trains'])

        wd[cnum] = dict(
            cars=cars, rev=rev, dh=dh,
            tot=tot if tot > 0 else "",
            start=layover, end=layover,
            first=first, final=final,
            next=na(r['next_cycle']),
            cmf=cmf, cmf_arr=cmf_arr, cmf_dep=cmf_dep,
            layover=layover,
            morning_trains=morning_raw,
            evening_trains=evening_raw,
        )

    # ── weekend (Sa / Su) ─────────────────────────────────────────────────────
    wknd = {}
    for _, r in df_cwk.iterrows():
        bc      = int(float(na(r['weekday_cycle'])))
        layover = clean(r['layover_location'])
        cars    = na(r['cars'])
        sa_rev  = to_int(r['sa_trip_count'])
        su_rev  = to_int(r['su_trip_count'])

        sa_trains = clean(r['saturday_trains'])
        su_trains = clean(r['sunday_trains'])

        wknd[bc] = {
            'Sa': dict(cars=cars, rev=sa_rev, dh=0,
                       tot=sa_rev if isinstance(sa_rev,int) else "",
                       start=layover, end=layover,
                       next=f"{bc}Su",
                       cmf="No", cmf_arr="N/A", cmf_dep="N/A",
                       layover=layover,
                       morning_trains=sa_trains,
                       evening_trains=""),
            'Su': dict(cars=cars, rev=su_rev, dh=0,
                       tot=su_rev if isinstance(su_rev,int) else "",
                       start=layover, end=layover,
                       next=str(bc),
                       cmf="No", cmf_arr="N/A", cmf_dep="N/A",
                       layover=layover,
                       morning_trains=su_trains,
                       evening_trains=""),
        }

    # Add "or XSa" suffix when next cycle has Saturday service
    for cnum, data in wd.items():
        nxt = data['next']
        if nxt == "": continue
        try:
            nxt_num = int(float(str(nxt)))
            if nxt_num in wknd:
                data['next'] = f"{nxt_num} or {nxt_num}Sa"
        except:
            pass

    return wd, wknd, train_routes, laus_dwell_lkp, olp_dwell_lkp

# ── move sequence builder ─────────────────────────────────────────────────────
def get_dwell(curr_tid, next_tid, end_station, laus_lkp, olp_lkp):
    """
    Look up dwell time between the end of curr_tid move and start of next_tid move.
    If end_station == LAUS → LAUS dwell lookup.
    Otherwise → OLP dwell lookup.
    """
    c = str(curr_tid); n = str(next_tid)
    if end_station == 'LAUS':
        return laus_lkp.get((c, n), '')
    else:
        return olp_lkp.get((c, n), '')

def build_cycle_moves(fields, train_routes, laus_lkp, olp_lkp):
    """
    Build a list of up to MAX_MOVES tuples: (train_id_str, start, end, dwell)
    for a weekday cycle.

    Two patterns:
      CMF-based   layover == 'CMF':
          first_morningQ CMF→LAUS | morning trains | evening trains | last_eveningQ LAUS→CMF

      OLP-based   layover != 'CMF':
        With midday CMF (cmf_arr and cmf_dep both set):
          morning trains | last_morningQ LAUS→CMF | first_eveningQ CMF→LAUS | evening trains
        Without CMF:
          morning trains | evening trains
    """
    morning = parse_trains(fields.get('morning_trains', ''))
    evening = parse_trains(fields.get('evening_trains', ''))
    layover = fields.get('layover', '')
    cmf_arr = clean(fields.get('cmf_arr', ''))
    cmf_dep = clean(fields.get('cmf_dep', ''))

    # Pure LAYOVER cycle (cycle 35) — no moves
    m_raw = clean(fields.get('morning_trains',''))
    if m_raw.upper() == 'LAYOVER':
        return []

    if not morning and not evening:
        return []

    is_cmf_based  = (layover == 'CMF')
    # Turn cycles: CMF is visited via an en-route turn move, not a separate Q deadhead.
    # We cannot automatically reconstruct the turn move, so skip explicit CMF deadheads.
    is_turn_cmf = (cmf_arr.startswith('Turn') or cmf_dep.startswith('Turn'))
    # Has midday CMF dwell: both times are real clock times (not N/A/empty/Turn)
    has_midday_cmf = (
        not is_cmf_based
        and not is_turn_cmf
        and bool(cmf_arr) and bool(cmf_dep)
        and cmf_arr not in ('N/A',) and cmf_dep not in ('N/A',)
    )
    # Also verify last morning train ends at LAUS (required for LAUS→CMF deadhead)
    last_morning_route = train_routes.get(morning[-1], ('','')) if morning else ('','')
    if has_midday_cmf and last_morning_route[1] != 'LAUS':
        has_midday_cmf = False  # can't do LAUS→CMF if last morning train doesn't arrive LAUS

    moves = []

    if is_cmf_based:
        # ── CMF-based: deadhead out at start, deadhead in at end ─────────────
        if morning:
            dh_out = f'{morning[0]}Q'
            laus_dwell_val = laus_lkp.get((dh_out, str(morning[0])), '')
            moves.append((dh_out, 'CMF', 'LAUS', laus_dwell_val))

        for i, tid in enumerate(morning):
            route = train_routes.get(tid, ('', ''))
            frm, to = route
            if i < len(morning) - 1:
                dwell = get_dwell(tid, morning[i+1], to, laus_lkp, olp_lkp)
            else:
                # Last morning → dwell before first evening
                dwell = get_dwell(tid, evening[0], to, laus_lkp, olp_lkp) if evening else 'N/A'
            moves.append((str(tid), frm, to, dwell))

        for i, tid in enumerate(evening):
            route = train_routes.get(tid, ('', ''))
            frm, to = route
            if i < len(evening) - 1:
                dwell = get_dwell(tid, evening[i+1], to, laus_lkp, olp_lkp)
            else:
                # Last evening → dwell at LAUS before CMF deadhead
                dh_in = f'{tid}Q'
                dwell = laus_lkp.get((str(tid), dh_in), '')
            moves.append((str(tid), frm, to, dwell))

        if evening:
            last_e = evening[-1]
            moves.append((f'{last_e}Q', 'LAUS', 'CMF', 'N/A'))

    else:
        # ── OLP-based ─────────────────────────────────────────────────────────
        for i, tid in enumerate(morning):
            route = train_routes.get(tid, ('', ''))
            frm, to = route
            if i < len(morning) - 1:
                dwell = get_dwell(tid, morning[i+1], to, laus_lkp, olp_lkp)
            elif has_midday_cmf:
                # Last morning → dwell at LAUS before CMF deadhead out
                dh_out = f'{tid}Q'
                dwell = laus_lkp.get((str(tid), dh_out), '')
            else:
                # No CMF → dwell before first evening
                dwell = get_dwell(tid, evening[0], to, laus_lkp, olp_lkp) if evening else 'N/A'
            moves.append((str(tid), frm, to, dwell))

        if has_midday_cmf and morning:
            last_m   = morning[-1]
            dh_out   = f'{last_m}Q'
            cmf_dwell = calc_time_diff(cmf_arr, cmf_dep)
            moves.append((dh_out, 'LAUS', 'CMF', cmf_dwell))

            if evening:
                first_e = evening[0]
                dh_in   = f'{first_e}Q'
                laus_after = laus_lkp.get((dh_in, str(first_e)), '')
                moves.append((dh_in, 'CMF', 'LAUS', laus_after))

        for i, tid in enumerate(evening):
            route = train_routes.get(tid, ('', ''))
            frm, to = route
            if i < len(evening) - 1:
                dwell = get_dwell(tid, evening[i+1], to, laus_lkp, olp_lkp)
            else:
                dwell = 'N/A'
            moves.append((str(tid), frm, to, dwell))

    return moves[:MAX_MOVES]

def get_fields(cid, wd, wknd):
    m2 = re.match(r'^(\d+)(Sa|Su)$', str(cid).strip())
    if m2:
        bc  = int(m2.group(1))
        var = m2.group(2)
        rec = wknd.get(bc, {}).get(var)
        if rec:
            rec = dict(rec)
            rec['preserve_times'] = True
        return rec
    else:
        try:
            cnum = int(float(str(cid).strip()))
            rec  = wd.get(cnum)
            if rec: rec = dict(rec); rec['preserve_times'] = False
            return rec
        except:
            return None

# ── styling helpers ───────────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F3864")
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
DATA_FONT   = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center")
TBD_FILL    = PatternFill("solid", fgColor="FFD966")
TURN_FILL   = PatternFill("solid", fgColor="FFE699")
SUSP_FILL   = PatternFill("solid", fgColor="F4CCCC")
WARN_FONT   = Font(name="Arial", size=10, bold=True, color="7F4F00")

def style_header_row(ws, n_cols):
    for c in range(1, n_cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.border = BORDER; cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

def style_data_row(ws, row_num, n_cols):
    fill = ALT_FILL if row_num % 2 == 0 else WHITE_FILL
    for c in range(1, n_cols+1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill; cell.font = DATA_FONT
        cell.border = BORDER; cell.alignment = CENTER
    ws.row_dimensions[row_num].height = 18

# ── write move sequence into a row ───────────────────────────────────────────
def write_moves(ws, row_num, moves):
    """Write up to MAX_MOVES move tuples into cols 8–43, blank unused slots."""
    for slot in range(MAX_MOVES):
        base = MOVE_BASE + slot * MOVE_STRIDE
        if slot < len(moves):
            tid, start, end, dwell = moves[slot]
            ws.cell(row=row_num, column=base).value     = tid   or None
            ws.cell(row=row_num, column=base+1).value   = start or None
            ws.cell(row=row_num, column=base+2).value   = end   or None
            ws.cell(row=row_num, column=base+3).value   = dwell or None
        else:
            for offset in range(4):
                ws.cell(row=row_num, column=base+offset).value = None

# ── main update logic ─────────────────────────────────────────────────────────
def populate(master_path, template_path, output_path):
    print("Loading master ...")
    wd, wknd, train_routes, laus_lkp, olp_lkp = build_lookups(master_path)
    print(f"  Weekday cycles       : {len(wd)}")
    print(f"  Weekend cycle groups : {len(wknd)}")
    print(f"  Train routes         : {len(train_routes)}")
    print(f"  LAUS dwell pairs     : {len(laus_lkp)}")
    print(f"  OLP dwell pairs      : {len(olp_lkp)}")

    print(f"Opening template: {template_path}")
    wb = load_workbook(template_path)
    ws = wb["Weekday"]
    n_cols = ws.max_column

    print(f"  Sheet columns   : {n_cols}")
    print(f"  Sheet data rows : {ws.max_row - 1}")

    style_header_row(ws, n_cols)

    updated = 0; skipped = 0

    for row_num in range(2, ws.max_row + 1):
        cid_cell = ws.cell(row=row_num, column=COL_CYCLE)
        cid = cid_cell.value
        if cid is None: continue

        fields = get_fields(cid, wd, wknd)

        if fields is None:
            # Suspended cycle — CLEAR all data cols so no stale old-schedule
            # trains survive in the output, then highlight the row red.
            for c in range(2, n_cols + 1):
                ws.cell(row=row_num, column=c).value = None
            style_data_row(ws, row_num, n_cols)
            for c in range(1, n_cols + 1):
                ws.cell(row=row_num, column=c).fill = SUSP_FILL
            ws.cell(row=row_num, column=COL_CYCLE).font = WARN_FONT
            skipped += 1
            continue

        preserve_times = fields.get('preserve_times', False)

        # ── header / summary columns ──────────────────────────────────────────
        # Compute Tot explicitly as an integer so we never rely on a template
        # formula cell (openpyxl strips formula caches on save → shows None).
        rev_val = fields['rev']
        dh_val  = fields['dh']
        tot_val = fields['tot']
        if tot_val == "" and isinstance(rev_val, int) and isinstance(dh_val, int):
            tot_val = rev_val + dh_val
        elif tot_val == "" and rev_val == 'TBD' and isinstance(dh_val, int):
            tot_val = dh_val  # TBD rev: show just DH count so cell isn't blank

        updates = {
            COL_CARS:    fields['cars'],
            COL_REV:     rev_val,
            COL_DH:      dh_val,
            COL_TOT:     tot_val,
            COL_START:   fields['start'],
            COL_END:     fields['end'],
            COL_NEXT:    fields['next'],
            COL_CMF:     fields['cmf'],
            COL_CMF_ARR: fields['cmf_arr'],
            COL_CMF_DEP: fields['cmf_dep'],
        }
        if not preserve_times:
            updates[COL_FIRST_MOVE] = fields['first']
            updates[COL_FINAL_MOVE] = fields['final']

        for col, val in updates.items():
            ws.cell(row=row_num, column=col).value = val if val != "" else None

        # ── move sequences (cols 8–43) ─────────────────────────────────────────
        moves = build_cycle_moves(fields, train_routes, laus_lkp, olp_lkp)
        write_moves(ws, row_num, moves)

        style_data_row(ws, row_num, n_cols)

        # Sketchy highlights
        if fields['rev'] == "TBD" or str(fields.get('rev','')).upper() == "TBD":
            for c in [COL_REV, COL_DH, COL_TOT]:
                cell = ws.cell(row=row_num, column=c)
                cell.fill = TBD_FILL; cell.font = WARN_FONT

        cmf_arr_val = str(fields.get('cmf_arr','') or '')
        cmf_dep_val = str(fields.get('cmf_dep','') or '')
        if cmf_arr_val.startswith('Turn') or cmf_dep_val.startswith('Turn'):
            for c in [COL_CMF, COL_CMF_ARR, COL_CMF_DEP]:
                cell = ws.cell(row=row_num, column=c)
                cell.fill = TURN_FILL; cell.font = WARN_FONT

        updated += 1

    col_widths = [14,10,11,16,10,18,12] + [13,14,12,14]*9 + [18,12,16,10,16,18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "H2"

    print("Populating Weekend sheet ...")
    populate_weekend(wb, wd, wknd, train_routes, laus_lkp, olp_lkp)

    wb.save(output_path)

    print(f"\nDone → {output_path}")
    print(f"  Rows updated     : {updated}")
    print(f"  Rows skipped     : {skipped}  (suspended/empty cycles)")

# ── validation ────────────────────────────────────────────────────────────────
def validate(output_path, master_path):
    print("\n── Data check ──────────────────────────────────────────────")
    xf = pd.ExcelFile(output_path)
    print(f"  Sheets: {xf.sheet_names}")

    df = pd.read_excel(output_path, sheet_name="Weekday")
    print(f"\n  Weekday: {len(df)} rows")
    print(f"    Rows with cars      : {df['Number of Cars'].notna().sum()}")
    print(f"    CMF = Yes           : {(df['CMF Train?']=='Yes').sum()}")
    print(f"    Rows with Train ID  : {df['Train ID'].notna().sum()}")

    # Spot-check a few cycles
    for chk_cycle in [2, 5, 7, 9, 22]:
        row = df[df['Current Day Cycle'] == chk_cycle]
        if row.empty: continue
        tid = row['Train ID'].values[0]
        s   = row['Start Location'].values[0]
        e   = row['End Location'].values[0]
        print(f"    Cycle {chk_cycle:2d} move 1: {tid} {s}→{e}")

    if "Weekend" in xf.sheet_names:
        dw = pd.read_excel(output_path, sheet_name="Weekend")
        print(f"\n  Weekend: {len(dw)} rows")
        print(f"    Rows with cars      : {dw['Number of Cars'].notna().sum()}")

    print("────────────────────────────────────────────────────────────")

# ═══════════════════════════════════════════════════════════════════════════════
# WEEKEND SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def populate_weekend(wb, wd, wknd, train_routes, laus_lkp, olp_lkp):
    ws_wd   = wb["Weekday"]
    ws_wknd = wb["Sheet1"]
    ws_wknd.title = "Weekend"
    n_cols  = ws_wd.max_column

    for c in range(1, n_cols + 1):
        src  = ws_wd.cell(row=1, column=c)
        dest = ws_wknd.cell(row=1, column=c)
        dest.value = src.value
        dest.fill  = HDR_FILL; dest.font = HDR_FONT
        dest.border = BORDER;  dest.alignment = CENTER
    ws_wknd.row_dimensions[1].height = 30

    sa_su_rows = []
    for row_num in range(2, ws_wd.max_row + 1):
        cid = ws_wd.cell(row=row_num, column=COL_CYCLE).value
        if cid is None: continue
        if re.search(r'(Sa|Su)$', str(cid).strip()):
            row_data = [ws_wd.cell(row=row_num, column=c).value
                        for c in range(1, n_cols + 1)]
            sa_su_rows.append(row_data)

    for dest_row, row_data in enumerate(sa_su_rows, 2):
        cid    = str(row_data[COL_CYCLE - 1]).strip()
        fields = get_fields(cid, wd, wknd)

        for c, val in enumerate(row_data, 1):
            ws_wknd.cell(row=dest_row, column=c).value = val

        if fields:
            # Compute Tot explicitly (same as weekday — avoid formula None)
            rv = fields['rev']; dv = fields['dh']; tv = fields['tot']
            if tv == "" and isinstance(rv, int) and isinstance(dv, int):
                tv = rv + dv
            # Update header/summary columns
            updates = {
                COL_CARS:    fields['cars'],
                COL_REV:     rv,
                COL_DH:      dv,
                COL_TOT:     tv,
                COL_START:   fields['start'],
                COL_END:     fields['end'],
                COL_NEXT:    fields['next'],
                COL_CMF:     fields['cmf'],
                COL_CMF_ARR: fields['cmf_arr'],
                COL_CMF_DEP: fields['cmf_dep'],
            }
            for col, val in updates.items():
                ws_wknd.cell(row=dest_row, column=col).value = val if val != "" else None

            # Rebuild move sequences for Sa/Su too
            moves = build_cycle_moves(fields, train_routes, laus_lkp, olp_lkp)
            write_moves(ws_wknd, dest_row, moves)

        style_data_row(ws_wknd, dest_row, n_cols)

        if fields and (fields.get('rev') == "TBD" or
                       str(fields.get('rev','')).upper() == "TBD"):
            for c in [COL_REV, COL_DH, COL_TOT]:
                cell = ws_wknd.cell(row=dest_row, column=c)
                cell.fill = TBD_FILL; cell.font = WARN_FONT

    col_widths = [14,10,11,16,10,18,12] + [13,14,12,14]*9 + [18,12,16,10,16,18]
    for i, w in enumerate(col_widths, 1):
        ws_wknd.column_dimensions[get_column_letter(i)].width = w
    ws_wknd.freeze_panes = "H2"

    print(f"  Weekend sheet: {len(sa_su_rows)} rows written")

# ── entry point ───────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--master",   required=True)
    p.add_argument("--template", required=True)
    p.add_argument("--output",   required=True)
    args = p.parse_args()
    populate(args.master, args.template, args.output)
    validate(args.output, args.master)

if __name__ == "__main__":
    main()

"""
build_master.py  —  Build my_master.xlsx from CYCLES source file.
Usage:  python build_master.py --cycles CYCLES.xlsx --output my_master.xlsx
Sheets: cycles_weekday | cycles_weekend | laus_dwell | olp_dwell | train_reference | arrow_service
"""
import argparse, re, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────
HDR = {"cycles_weekday":"1F3864","cycles_weekend":"4D1F1F","laus_dwell":"0F4D3F",
       "olp_dwell":"3D1F64","train_reference":"4D2D0F","arrow_service":"1A4D2E"}
ALT = {"cycles_weekday":"EEF3FA","cycles_weekend":"FAF0EE","laus_dwell":"EDF6F2",
       "olp_dwell":"F3EEF9","train_reference":"FAF2EE","arrow_service":"EEF6F0"}

# ── helpers ───────────────────────────────────────────────────────────────────
def ft(val):
    if val is None: return ""
    if isinstance(val, datetime.time):      return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):  return val.strftime("%H:%M")
    s = str(val).strip(); return s if s.lower() not in ("none","nan") else ""

def fd(val):
    if val is None: return ""
    if isinstance(val, datetime.time):     h,m = val.hour, val.minute
    elif isinstance(val, datetime.datetime): h,m = val.hour, val.minute
    else: return str(val)
    if h==0 and m==0: return ""
    if h==0: return f"{m} min"
    if m==0: return f"{h} hr"
    return f"{h} hr {m} min"

def cv(v):
    if v is None: return ""
    if isinstance(v,float) and v==int(v): return int(v)
    return v

def clean_trains(raw):
    if raw is None: return ""
    s = str(raw)
    if s.upper() in ("LAYOVER","XXX"): return s.upper()
    if s.lower() in ("none","nan"): return ""
    s = re.sub(r'[….]+','',s)
    s = re.sub(r'Turns?\s+(to|from)\s+\d+','',s,flags=re.IGNORECASE)
    ids=[t.strip() for t in re.split(r'[\s\-\u2013\u2014,/]+',s.strip()) if t.strip().isdigit()]
    return ", ".join(ids) if ids else str(raw).strip()

def write_sheet(wb, title, headers, rows, widths, key):
    ws = wb.create_sheet(title=title)
    thin = Side(style="thin",color="CCCCCC")
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
    hf   = PatternFill("solid",fgColor=HDR[key])
    hfont= Font(name="Arial",bold=True,color="FFFFFF",size=10)
    df   = Font(name="Arial",size=10)
    ac   = Alignment(horizontal="center",vertical="center")
    af   = PatternFill("solid",fgColor=ALT[key])
    wf   = PatternFill("solid",fgColor="FFFFFF")
    for c,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=c,value=h)
        cell.fill=hf; cell.font=hfont; cell.border=bdr; cell.alignment=ac
    ws.row_dimensions[1].height=28
    for ri,row in enumerate(rows,2):
        fill=af if ri%2==0 else wf
        for ci,v in enumerate(row,1):
            cell=ws.cell(row=ri,column=ci,value=v)
            cell.fill=fill; cell.font=df; cell.border=bdr; cell.alignment=ac
        ws.row_dimensions[ri].height=16
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"
    print(f"  {title}: {len(rows)} rows")

# ═══════════════════════════════════════════════════════════════════════════════
# 1. CYCLES WEEKDAY
# ═══════════════════════════════════════════════════════════════════════════════
CW_H = ["cycle","cars","depart_cmf","evening_trains","arrive_layover","layover_location",
        "depart_layover","morning_trains","arrive_cmf","next_cycle","miles","weekday_trips",
        "sa_trip_count","su_trip_count","sa_trains","su_trains",
        "cmf_arrive_time","cmf_depart_time","cmf_dwell","notes"]
CW_W = [8,7,13,28,14,16,14,28,14,11,9,13,12,12,30,30,15,15,12,20]

def extract_cycles_weekday(wb):
    ws=wb["COMPLETE CYCLES"]
    cycles={}; order=[]
    current=None
    for row in ws.iter_rows(min_row=2,values_only=True):
        a=row[0]
        if a is None:
            if current is None: continue
            d,h2,n,o=row[3],row[7],row[13],row[14]
            if d and "SATURDAY" in str(d).upper():
                sa=re.sub(r'SATURDAY\s*','',str(d),flags=re.IGNORECASE).strip()
                cycles[current]["sa_trains"]=clean_trains(sa)
            if h2 and "SUNDAY" in str(h2).upper():
                su=re.sub(r'SUNDAY\s*','',str(h2),flags=re.IGNORECASE).strip()
                cycles[current]["su_trains"]=clean_trains(su)
            if n is not None and cycles[current]["sa_trip_count"]=="": cycles[current]["sa_trip_count"]=cv(n)
            if o is not None and cycles[current]["su_trip_count"]=="": cycles[current]["su_trip_count"]=cv(o)
            continue
        if isinstance(a,str) and not re.match(r'^\d+$',a.strip()):
            current=None; continue
        try: cnum=int(a)
        except: continue
        if cnum>40: break
        dc=row[2]; ac=row[8]; notes=[]
        if isinstance(dc,str) and "turn" in dc.lower(): notes.append(f"Depart: {dc.strip()}"); dc=None
        if isinstance(ac,str) and "turn" in ac.lower(): notes.append(f"Arrive: {ac.strip()}"); ac=None
        miles=row[10]; trips=row[12]
        current=cnum; order.append(cnum)
        cycles[cnum]={"cars":cv(row[1]),"depart_cmf":ft(dc),"evening_trains":clean_trains(row[3]),
            "arrive_layover":ft(row[4]),"layover":str(row[5]).strip() if row[5] else "",
            "depart_layover":ft(row[6]),"morning_trains":clean_trains(row[7]),"arrive_cmf":ft(ac),
            "next_cycle":cv(row[9]),"miles":round(float(miles),2) if isinstance(miles,(int,float)) else "",
            "weekday_trips":cv(trips) if str(trips) not in ("XXX","None","") else "TBD",
            "sa_trip_count":"","su_trip_count":"","sa_trains":"","su_trains":"","notes":"; ".join(notes)}
    return cycles, order

def extract_cmf_dwell(wb):
    ws=wb["CMF-EMF DWELL"]
    arr={}; dep={}; dwell={}
    in_cmf=True
    for row in ws.iter_rows(min_row=2,values_only=True):
        a=row[0]
        if isinstance(a,str) and "DWELL" in a.upper(): in_cmf="CMF" in a.upper(); continue
        if not in_cmf: continue
        if a is not None and str(a).strip() not in ("CYCLE",""):
            try:
                c=int(a); t=ft(row[2])
                if t:
                    # "Turns to/from XXX" means cycle visits CMF via a turn move —
                    # store as "Turn:XXX" so downstream code knows CMF=Yes but no fixed time
                    if isinstance(row[2],str) and "turn" in row[2].lower():
                        m=re.search(r'\d+',str(row[2]))
                        arr[c]="Turn:"+m.group() if m else "Turn"
                    else:
                        arr[c]=t
                if len(row)>8 and row[8]: dwell[c]=fd(row[8])
            except: pass
        if len(row)>4 and row[4] is not None and str(row[4]).strip() not in ("CYCLE",""):
            try:
                c=int(row[4]); t=ft(row[6])
                if t:
                    if isinstance(row[6],str) and "turn" in row[6].lower():
                        m=re.search(r'\d+',str(row[6]))
                        dep[c]="Turn:"+m.group() if m else "Turn"
                    else:
                        dep[c]=t
            except: pass
    return arr, dep, dwell

def build_cycles_weekday_rows(wb):
    cycles, order = extract_cycles_weekday(wb)
    arr, dep, dwell = extract_cmf_dwell(wb)
    rows=[]
    for cnum in order:
        c=cycles[cnum]
        rows.append([cnum,c["cars"],c["depart_cmf"],c["evening_trains"],c["arrive_layover"],
            c["layover"],c["depart_layover"],c["morning_trains"],c["arrive_cmf"],c["next_cycle"],
            c["miles"],c["weekday_trips"],c["sa_trip_count"],c["su_trip_count"],
            c["sa_trains"],c["su_trains"],arr.get(cnum,""),dep.get(cnum,""),dwell.get(cnum,""),c["notes"]])
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# 2. CYCLES WEEKEND
# ═══════════════════════════════════════════════════════════════════════════════
CWKND_H = ["weekday_cycle","cars","saturday_trains","sunday_trains","sa_trip_count","su_trip_count",
           "layover_location","total_trips","total_miles","fuel_gals","def_gals","notes"]
CWKND_W = [14,7,36,36,13,13,18,11,12,11,11,25]

def build_cycles_weekend_rows(wb):
    # Weekend groups from Weekends T M G L
    ws=wb["Weekends T M G L"]
    groups=[]; current=None; mrow=0
    for row in ws.iter_rows(min_row=1,max_row=75,values_only=True):
        a=row[0]; h=row[7] if len(row)>7 else None
        if a is not None and isinstance(a,(int,float)) and 1<=int(a)<=9:
            trains=[int(row[j]) for j in range(1,7) if len(row)>j and row[j] is not None and isinstance(row[j],(int,float))]
            layover=str(row[6]).strip() if len(row)>6 and row[6] and not isinstance(row[6],(int,float)) else ""
            trips=int(h) if isinstance(h,(int,float)) else None
            current={"cars":int(a),"trains":trains,"layover":layover,"trips":trips,"miles":None,"fuel":None,"def_g":None}
            groups.append(current); mrow=0
        elif current is not None and a is None and h is not None and isinstance(h,(int,float)):
            mrow+=1
            if mrow==1:   current["miles"]=round(float(h),2)
            elif mrow==2: current["fuel"] =round(float(h),2)
            elif mrow==3: current["def_g"]=round(float(h),4)

    # Weekend cycle mapping from COMPLETE CYCLES
    ws2=wb["COMPLETE CYCLES"]; wmap={}; cur=None
    for row in ws2.iter_rows(min_row=2,values_only=True):
        a=row[0]
        if a is not None:
            try: cur=int(a) if int(a)<=40 else None
            except: cur=None
            continue
        if cur is None: continue
        d,h2,n,o=row[3],row[7],row[13],row[14]
        sa_ids=[]; su_ids=[]
        if d and "SATURDAY" in str(d).upper():
            raw=re.sub(r'SATURDAY\s*','',str(d),flags=re.IGNORECASE).strip()
            sa_ids=[int(t) for t in re.split(r'[\s\-,]+',raw) if t.strip().isdigit()]
        if h2 and "SUNDAY" in str(h2).upper():
            raw=re.sub(r'SUNDAY\s*','',str(h2),flags=re.IGNORECASE).strip()
            su_ids=[int(t) for t in re.split(r'[\s\-,]+',raw) if t.strip().isdigit()]
        if sa_ids or su_ids:
            if cur not in wmap: wmap[cur]={"sa":[],"su":[],"sa_cnt":None,"su_cnt":None}
            if sa_ids: wmap[cur]["sa"]=sa_ids
            if su_ids: wmap[cur]["su"]=su_ids
            if n is not None: wmap[cur]["sa_cnt"]=cv(n)
            if o is not None: wmap[cur]["su_cnt"]=cv(o)

    def match(ids):
        for g in groups:
            if set(ids)&set(g["trains"]): return g
        return None

    rows=[]
    for cnum in sorted(wmap.keys()):
        w=wmap[cnum]; sa=w["sa"]; su=w["su"]
        g=match(sa) or match(su)
        notes="SA and SU trains differ" if sa!=su else ""
        rows.append([cnum,g["cars"] if g else "",
            ", ".join(str(t) for t in sa),", ".join(str(t) for t in su),
            w["sa_cnt"],w["su_cnt"],g["layover"] if g else "",
            g["trips"] if g else "",g["miles"] if g else "",
            g["fuel"] if g else "",g["def_g"] if g else "",notes])
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# 3. LAUS DWELL
# ═══════════════════════════════════════════════════════════════════════════════
LD_H = ["ib_train","ib_arrive","ib_from_cmf","ob_train","ob_depart","ob_to_cmf","laus_track","dwell","day_type"]
LD_W = [12,12,12,12,12,12,12,14,16]

def build_laus_dwell_rows(wb):
    ws=wb["LAUS DWELL"]; rows=[]
    for i,row in enumerate(ws.iter_rows(min_row=3,values_only=True),3):
        ib=row[0]; ob=row[4] if len(row)>4 else None
        if ib is None or str(ib) in ('I/B SYM','LAUS - WEEKDAY','LAUS - WEEKEND / HOLIDAY'): continue
        rows.append([str(ib).strip(),ft(row[1]),
            "Yes" if row[2] and "from cmf" in str(row[2]).lower() else "",
            str(ob).strip() if ob else "",
            ft(row[5]) if len(row)>5 else "",
            "Yes" if len(row)>6 and row[6] and "to cmf" in str(row[6]).lower() else "",
            str(row[3]).strip() if len(row)>3 and row[3] else "",
            fd(row[7]) if len(row)>7 else "","Weekday"])
        if len(row)>9 and row[9] is not None:
            ib2=row[9]; ob2=row[13] if len(row)>13 else None
            if str(ib2) not in ('I/B SYM',):
                rows.append([str(ib2).strip(),ft(row[10]) if len(row)>10 else "",
                    "Yes" if len(row)>11 and row[11] and "from cmf" in str(row[11]).lower() else "",
                    str(ob2).strip() if ob2 else "",
                    ft(row[14]) if len(row)>14 else "",
                    "Yes" if len(row)>15 and row[15] and "to cmf" in str(row[15]).lower() else "",
                    str(row[12]).strip() if len(row)>12 and row[12] else "",
                    fd(row[16]) if len(row)>16 else "","Weekend/Holiday"])
    return [r for r in rows if r[0] and r[0] not in ("I/B SYM",)]

# ═══════════════════════════════════════════════════════════════════════════════
# 4. OLP DWELL
# ═══════════════════════════════════════════════════════════════════════════════
OD_H = ["location","day_type","ob_train","ob_arrive","ib_train","ib_depart","dwell","notes"]
OD_W = [20,16,12,12,12,12,14,20]

def build_olp_dwell_rows(wb):
    ws=wb["OLP DWELL"]; rows=[]; location=""
    lmap={"EMF":"EMF","PERRIS":"SPS","LANCASTER":"LCS","EAST VENTURA":"EVC",
          "STUART MESA":"SMMF","RIVERSIDE":"RVS","MOORPARK":"MPK","AMF":"AMF"}
    for row in ws.iter_rows(min_row=1,values_only=True):
        a=row[0]
        if a is None: continue
        a_str=str(a).strip().upper()
        for k,v in lmap.items():
            if k in a_str: location=v; break
        if a_str in ("O/B SYM","I/B SYM"): continue
        if a and location:
            ob=str(a).strip(); ib=str(row[4]).strip() if len(row)>4 and row[4] else ""
            note=""
            if len(row)>2 and row[2]: note=str(row[2]).strip()
            if len(row)>6 and row[6]: note=(note+" "+str(row[6]).strip()).strip()
            wkday="Weekend/Holiday" if any(x in a_str for x in ("WEEKEND","HOLIDAY","SATURDAY","SUNDAY")) else "Weekday"
            rows.append([location,wkday,ob,ft(row[1]) if len(row)>1 else "",
                ib,ft(row[5]) if len(row)>5 else "",
                fd(row[7]) if len(row)>7 else "",note])
        if len(row)>9 and row[9] is not None and location:
            ob2=str(row[9]).strip(); ib2=str(row[13]).strip() if len(row)>13 and row[13] else ""
            note2=""
            if len(row)>11 and row[11]: note2=str(row[11]).strip()
            if len(row)>15 and row[15]: note2=(note2+" "+str(row[15]).strip()).strip()
            rows.append([location,"Weekend/Holiday",ob2,ft(row[10]) if len(row)>10 else "",
                ib2,ft(row[14]) if len(row)>14 else "",
                fd(row[16]) if len(row)>16 else "",note2])
    return [r for r in rows if r[2] and r[2] not in ("","O/B SYM")]

# ═══════════════════════════════════════════════════════════════════════════════
# 5. TRAIN REFERENCE  (weekday + weekend trains)
# ═══════════════════════════════════════════════════════════════════════════════
TR_H = ["train_id","service_type","from_station","to_station","miles",
        "median_capacity","median_peak_load","load_pct","avg_bicycles"]
TR_W = [10,14,15,15,9,16,16,10,14]

def extract_1xxx_weekend_trains(wb):
    """
    Dynamically extract all 1xxx weekend train IDs from LAUS DWELL + Weekends T M G L.
    Returns dict: train_id → (from_station, to_station, miles)
    """
    # ── Step 1: LAUS DWELL — direction (I/B=arrives LAUS, O/B=departs LAUS) ──
    ws_ld = wb["LAUS DWELL"]
    ib_set = set(); ob_set = set()
    cmf_from = set()  # trains that have "From CMF" flag → far end is CMF not layover
    cmf_to   = set()  # trains that have "To CMF"   flag → far end is CMF not layover
    def _clean_id(v):
        if v is None: return None
        s=str(v).strip().replace('Q','')
        return int(s) if s.isdigit() else None
    for row in ws_ld.iter_rows(min_row=3,values_only=True):
        ib=_clean_id(row[9]);  ob=_clean_id(row[13])
        ib_info=str(row[11]).lower() if len(row)>11 and row[11] else ""
        ob_info=str(row[15]).lower() if len(row)>15 and row[15] else ""
        if ib and ib>=1000:
            ib_set.add(ib)
            if "from cmf" in ib_info: cmf_from.add(ib)
        if ob and ob>=1000:
            ob_set.add(ob)
            if "to cmf" in ob_info: cmf_to.add(ob)

    # ── Step 2: Weekends T M G L — per-train miles + layover location ─────────
    ws_wt = wb["Weekends T M G L"]
    train_miles   = {}  # train_id → miles
    train_layover = {}  # train_id → layover code
    prev_trains = []; prev_layover = ""
    for row in ws_wt.iter_rows(min_row=1,max_row=75,values_only=True):
        a=row[0]
        if a is not None and isinstance(a,(int,float)) and 1<=int(a)<=9:
            prev_trains=[row[j] for j in range(1,7) if len(row)>j
                         and row[j] is not None and isinstance(row[j],(int,float))]
            prev_layover=str(row[6]).strip() if len(row)>6 and row[6] and not isinstance(row[6],(int,float)) else ""
        elif prev_trains and row[0] is None:
            miles_vals=[row[j] for j in range(1,7) if len(row)>j
                        and row[j] is not None and isinstance(row[j],(int,float))]
            if miles_vals and any(10<m<200 for m in miles_vals):
                for tid,mi in zip(prev_trains,miles_vals):
                    if int(tid)>=1000:
                        train_miles[int(tid)]=round(float(mi),2)
                        train_layover[int(tid)]=prev_layover
                prev_trains=[]

    # ── Step 3: Layover code → station abbreviation ───────────────────────────
    def layover_to_station(loc, tid, miles):
        loc=loc.upper()
        # Trains ≤38 miles in an LCS group are actually CMF-segment only
        if miles and miles<=38 and any(x in loc for x in ("LCS","CMF")):
            return "CMF"
        if "EVC" in loc:  return "EVC"
        if "LCS" in loc:  return "LCS"
        if "CMF" in loc:  return "CMF"
        if "SNB" in loc:  return "SNB"
        if "OSD" in loc:  return "OSD"
        if "SPS" in loc:  return "SPS"
        if "SMMF" in loc: return "SMMF"
        if "EMF" in loc:  return "EMF"
        if "RVS" in loc:  return "RVS"
        return loc.split("/")[0].strip()

    # ── Step 4: Combine ───────────────────────────────────────────────────────
    all_1xxx = ib_set | ob_set | set(train_miles.keys())
    result={}
    for tid in sorted(all_1xxx):
        if tid<1000: continue
        miles = train_miles.get(tid,"")
        loc   = train_layover.get(tid,"")
        far   = layover_to_station(loc,tid,miles)
        is_ib = tid in ib_set
        is_ob = tid in ob_set
        if is_ib and not is_ob:   frm,to = far,"LAUS"
        elif is_ob and not is_ib: frm,to = "LAUS",far
        else:                     frm,to = far,"LAUS"  # turn train: treat as IB
        result[tid]=(frm,to,miles)

    # ── Step 5: trains 1857-1860 (OSD↔SNB, don't stop at LAUS) ──────────────
    for tid,mi in [(1857,101.2),(1858,101.2),(1859,101.2),(1860,101.2)]:
        if tid not in result:
            # Odd → OSD→SNB direction; Even → SNB→OSD (Metrolink convention)
            frm,to = ("OSD","SNB") if tid%2==1 else ("SNB","OSD")
            result[tid]=(frm,to,mi)

    return result

def build_train_reference_rows(wb):
    ws=wb["LoadFactors"]; rows=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if row[0] is None or str(row[0]) in ("TRAIN",): continue
        try: tid=int(row[0])
        except: continue
        lp=row[3]
        rows.append([tid,"Weekday",
            str(row[5]).strip() if row[5] else "",
            str(row[6]).strip() if row[6] else "",
            round(float(row[7]),2) if isinstance(row[7],(int,float)) else "",
            cv(row[1]),cv(row[2]),
            f"{round(float(lp)*100,1)}%" if isinstance(lp,float) else "",
            round(float(row[4]),1) if isinstance(row[4],float) else cv(row[4])])
    # Add 1xxx weekend trains (dynamically extracted from LAUS DWELL + WTMGL)
    existing_ids={r[0] for r in rows}
    weekend_1xxx = extract_1xxx_weekend_trains(wb)
    for tid,(frm,to,miles) in sorted(weekend_1xxx.items()):
        if tid not in existing_ids:
            rows.append([tid,"Weekend/Holiday",frm,to,miles,"","","",""])
    rows.sort(key=lambda r:(r[1]!="Weekday", r[0]))
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# 6. ARROW SERVICE
# ═══════════════════════════════════════════════════════════════════════════════
AR_H = ["run","crew","dmu_units","service_type","departure","turns","arrival","next_run","miles","trips"]
AR_W = [7,7,10,16,12,60,12,12,9,8]

def build_arrow_rows(wb):
    ws=wb["ARROW Cycles"]; rows=[]; stype="Weekday"
    for i,row in enumerate(ws.iter_rows(min_row=1,values_only=True),1):
        if i==1: continue
        a=row[0]
        if a is None and row[4] is not None:
            if "WEEKEND" in str(row[4]).upper() or "HOLIDAY" in str(row[4]).upper(): stype="Weekend/Holiday"
            continue
        if isinstance(a,str) and a.upper() in ("RUN",): continue
        run=cv(row[0]) if row[0] is not None else ""
        crew=str(row[1]).strip() if row[1] else ""
        if not crew and not run: continue
        turns=str(row[4]).strip() if row[4] else ""
        if not (turns and re.search(r'\d{4}',turns)): continue
        rows.append([run,crew,cv(row[2]) if row[2] is not None else "",stype,
            ft(row[3]) if row[3] else "",turns,
            ft(row[5]) if len(row)>5 and row[5] else "",
            str(row[6]).strip() if len(row)>6 and row[6] else "",
            round(float(row[7]),2) if len(row)>7 and isinstance(row[7],(int,float)) else "",
            cv(row[9]) if len(row)>9 and row[9] is not None else ""])
    return rows

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    parser=argparse.ArgumentParser()
    parser.add_argument("--cycles",required=True)
    parser.add_argument("--output",required=True)
    args=parser.parse_args()
    print(f"Reading {args.cycles} ...")
    wb=load_workbook(args.cycles,data_only=True)
    print("Building sheets ...")
    wb_out=Workbook(); wb_out.remove(wb_out.active)
    write_sheet(wb_out,"cycles_weekday", CW_H,    build_cycles_weekday_rows(wb), CW_W,    "cycles_weekday")
    write_sheet(wb_out,"cycles_weekend", CWKND_H, build_cycles_weekend_rows(wb), CWKND_W, "cycles_weekend")
    write_sheet(wb_out,"laus_dwell",     LD_H,    build_laus_dwell_rows(wb),     LD_W,    "laus_dwell")
    write_sheet(wb_out,"olp_dwell",      OD_H,    build_olp_dwell_rows(wb),      OD_W,    "olp_dwell")
    write_sheet(wb_out,"train_reference",TR_H,    build_train_reference_rows(wb),TR_W,    "train_reference")
    write_sheet(wb_out,"arrow_service",  AR_H,    build_arrow_rows(wb),          AR_W,    "arrow_service")
    wb_out.save(args.output)
    print(f"\nDone → {args.output}")

if __name__=="__main__":
    main()
